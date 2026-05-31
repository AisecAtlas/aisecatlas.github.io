"""AISec Atlas POC 평가표 빌더 (CSV → 스타일 xlsx).

myai/knowledge/iam-pam-evaluation/build_xlsx.py 패턴 이식 — 단, 항목을 하드코딩하지 않고
`_pipeline/poc/<slug>/eval.csv` 를 읽어 워크북을 생성한다. aisec-poc-designer 가 솔루션마다
eval.csv 를 만들고 이 스크립트로 xlsx 를 빌드한다.

사용법:
    python build_eval_xlsx.py <slug>
    예) python build_eval_xlsx.py crowdstrike-charlotte
    → 입력 : _pipeline/poc/<slug>/eval.csv  (UTF-8, BOM 허용)
    → 출력 : _pipeline/poc/<slug>/eval.xlsx

eval.csv 스키마(헤더 고정, 순서 무관·이름 기준):
    대분류,ID,평가항목,L0기준,L1기준,L2기준,L3기준,가중치,증빙요청자료,점수,근거표준
  - 가중치: 1~5 (개보법/필수 통제는 높게)
  - 점수  : 공란(사람이 시연받고 0~3 입력) — H2
"""
import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

HERE = os.path.dirname(os.path.abspath(__file__))

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
DOMAIN_FILL = PatternFill("solid", fgColor="D9E1F2")
DOMAIN_FONT = Font(bold=True, size=11)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

COLUMNS = ["대분류", "ID", "평가항목", "L0기준", "L1기준", "L2기준", "L3기준",
           "가중치", "증빙요청자료", "점수", "근거표준"]
WIDTHS = [22, 9, 28, 26, 26, 26, 26, 9, 30, 8, 22]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER


def style_row(ws, row, ncols, is_domain=False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment, cell.border = WRAP, BORDER
        if is_domain:
            cell.fill, cell.font = DOMAIN_FILL, DOMAIN_FONT


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def read_rows(csv_path):
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        missing = [c for c in COLUMNS if c not in reader.fieldnames]
        if missing:
            raise SystemExit(f"eval.csv 헤더 누락: {missing}\n필요: {COLUMNS}")
        return [r for r in reader]


def build(slug):
    poc_dir = os.path.join(HERE, "poc", slug)
    csv_path = os.path.join(poc_dir, "eval.csv")
    out_path = os.path.join(poc_dir, "eval.xlsx")
    if not os.path.exists(csv_path):
        raise SystemExit(f"입력 없음: {csv_path}")

    rows = read_rows(csv_path)
    wb = Workbook()

    # ── 사용방법 시트 ──
    ws0 = wb.active
    ws0.title = "사용방법"
    set_widths(ws0, [4, 26, 80])
    intro = [
        ("", f"AISec Atlas POC 평가표 — {slug}", ""),
        ("", "채점 척도", ""),
        ("", "Level 0", "기능 부재 / N/A"),
        ("", "Level 1 (기본)", "기능은 있으나 수동·제한적·기본 설정 수준"),
        ("", "Level 2 (표준)", "표준화된 기능, 정책 자동 적용"),
        ("", "Level 3 (고급)", "자동화·세분화·증적 자동, 모범사례(BP) 수준"),
        ("", "", ""),
        ("", "채점 절차", ""),
        ("", "1", "각 벤더에 '증빙요청자료' 컬럼을 RFP 형태로 전달"),
        ("", "2", "데모/POC 환경에서 직접 시연받음 → '점수' 컬럼에 0~3 입력 (사람)"),
        ("", "3", "평가자 2인 이상 교차검증, 차이 1 이상이면 회의로 조정"),
        ("", "", ""),
        ("", "주의", "L0~L3 기준 문구를 그대로 판정 기준으로 사용(주관 배제). '점수'는 비워둔 상태로 배포됨."),
    ]
    for i, (a, b, c) in enumerate(intro, start=1):
        ws0.cell(row=i, column=1, value=a)
        ws0.cell(row=i, column=2, value=b).alignment = WRAP
        ws0.cell(row=i, column=3, value=c).alignment = WRAP
    ws0.cell(row=1, column=2).font = Font(bold=True, size=15)
    for r in (2, 8):
        ws0.cell(row=r, column=2).font = Font(bold=True, size=12, color="1F4E78")

    # ── 평가표 시트 ──
    ws = wb.create_sheet("PoC_평가표")
    ws.append(COLUMNS)
    style_header(ws, 1, len(COLUMNS))
    set_widths(ws, WIDTHS)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    score_col = COLUMNS.index("점수") + 1       # J
    weight_col = COLUMNS.index("가중치") + 1     # H
    current_domain = None
    r = 2
    first_data = None
    item_rows = []                               # 실제 평가항목 행(합성 대분류 헤더 제외)
    for row in rows:
        dom = (row.get("대분류") or "").strip()
        if dom and dom != current_domain:
            ws.cell(row=r, column=1, value=dom)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLUMNS))
            style_row(ws, r, len(COLUMNS), is_domain=True)
            current_domain = dom
            r += 1
        for ci, col in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=ci, value=(row.get(col) or "").strip())
        style_row(ws, r, len(COLUMNS))
        ws.row_dimensions[r].height = 46
        if first_data is None:
            first_data = r
        item_rows.append(r)
        r += 1
    last = r - 1
    item_count = len(item_rows)                  # = len(rows)

    wl, sl = get_column_letter(weight_col), get_column_letter(score_col)
    if first_data:
        ws.conditional_formatting.add(
            f"{sl}{first_data}:{sl}{last}",
            ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                           mid_type='num', mid_value=1.5, mid_color='FFEB84',
                           end_type='num', end_value=3, end_color='63BE7B'))

    # ── 점수계산 시트 ──
    wsC = wb.create_sheet("점수계산")
    set_widths(wsC, [28, 18, 18, 40])
    wsC.cell(row=1, column=1, value="항목").font = Font(bold=True)
    wsC.cell(row=1, column=2, value="값").font = Font(bold=True)
    sheet = "'PoC_평가표'"
    calc = [
        ("가중합 (Σ 가중치×점수)",
         f"=SUMPRODUCT({sheet}!{wl}{first_data}:{wl}{last},{sheet}!{sl}{first_data}:{sl}{last})"),
        ("만점 (Σ 가중치×3)",
         f"=SUMPRODUCT({sheet}!{wl}{first_data}:{wl}{last})*3"),
        ("달성률(%)", "=IFERROR(B2/B3*100,0)"),
        ("채점 완료 항목 수", f"=COUNT({sheet}!{sl}{first_data}:{sl}{last})"),
        ("전체 항목 수", f"={item_count}"),
    ]
    for i, (label, formula) in enumerate(calc, start=2):
        wsC.cell(row=i, column=1, value=label).alignment = WRAP
        wsC.cell(row=i, column=2, value=formula)
    wsC.cell(row=8, column=1, value="대분류별 부분점수가 필요하면 SUMPRODUCT + 대분류 필터로 계산").alignment = WRAP

    wb.save(out_path)
    print(f"Saved: {out_path}  (항목 {item_count}개)")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("usage: python build_eval_xlsx.py <slug>")
    build(sys.argv[1])
